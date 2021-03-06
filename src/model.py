#!/usr/bin/env python
# -*- coding: latin-1; py-indent-offset:4 -*-
################################################################################
# 
#   Copyright (C) 2014 Daniel Rodriguez
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
from collections import OrderedDict
import copy
import datetime
import itertools
from operator import attrgetter
import os
import os.path
import shutil
import threading

import openpyxl
import pythoncom
from win32com.client import Dispatch, DispatchEx
import win32com.client.gencache
import xlrd
import xlsxwriter

from config import ConfigInt, ConfigPrefix, ConfigString, ConfigBool
from excelmodel import ExcelInput
from mvcbase import ModelRole, PubSend
from rpcinterface import RpcSupport
from rpcsupport import RpcInterface
from tcmodel import TestCatalog, TestCase, TestCaseInPlan, Ticket, TicketChangeLog

@ModelRole
class Model(object):
    winwidth = ConfigInt(name='winwidth', defvalue=-1)
    winheight = ConfigInt(name='winheight', defvalue=-1)

    serverurl = ConfigString(name='serverurl', defvalue='')
    serverusername = ConfigString(name='server', defvalue='')
    serverpassword = ConfigString(name='serverpassword', defvalue='')

    uploadexcel = ConfigString(name='uploadexcel', defvalue='')
    uploadcatalogname = ConfigString(name='uploadcatalogname', defvalue='')
    uploadautosheets = ConfigBool(name='uploadautosheets', defvalue=False)
    uploadautocatalogs = ConfigBool(name='uploadautocatalogs', defvalue=False)

    uploadusecatalog = ConfigBool(name='uploadusecatalog', defvalue=False)
    uploadupdatetc = ConfigBool(name='uploadupdatetc', defvalue=False)

    downloadexcel = ConfigString(name='downloadexcel', defvalue='')

    lifecardexcel = ConfigString(name='lifecardexcel', defvalue='')
    lifecardattachdir = ConfigString(name='lifecardattachdir', defvalue='')
    lcdownexcelnotsave = ConfigBool(name='lcdownexcelnotsave', defvalue=False)
    lcdownkeepexcelopen = ConfigBool(name='lcdownkeepexcelopen', defvalue=False)
    lcdownmakecopy = ConfigBool(name='lcdownmakecopy', defvalue=False)
    lcdownfiltervendorcomments = ConfigBool(name='lcdownfiltervendorcomments', defvalue=False)
    lcdowncopywithvendorcomments = ConfigBool(name='lcdowncopywithvendorcomments', defvalue=False)
    lcdownopen = ConfigBool(name='lcdownopen', defvalue=True)
    lcdownclosed = ConfigBool(name='lcdownclosed', defvalue=True)
    lcdownfixed = ConfigBool(name='lcdownfixed', defvalue=True)
    lcdowninvest = ConfigBool(name='lcdowninvest', defvalue=True)
    lcdownreject = ConfigBool(name='lcdownreject', defvalue=True)
    lcdownnew  = ConfigBool(name='lcdownnew', defvalue=False)

    lcdownattachopen = ConfigBool(name='lcdownattachopen', defvalue=True)
    lcdownattachclosed = ConfigBool(name='lcdownattachclosed', defvalue=True)
    lcdownattachinvestigation = ConfigBool(name='lcdownattachinvestigation', defvalue=True)
    lcdownattachfixed = ConfigBool(name='lcdownattachfixed', defvalue=True)
    lcdownattachrejected = ConfigBool(name='lcdownattachrejected', defvalue=True)
    lcdownattachnew = ConfigBool(name='lcdownattachnew', defvalue=False)

    lifecardexcelup = ConfigString(name='lifecardexcelup', defvalue='')
    lifecardauthorup = ConfigString(name='lifecardauthorup', defvalue='')
    lcardup_sheetname = ConfigString(name='lcardup_sheetname', defvalue='Sheet Name')
    lcardup_row_start = ConfigString(name='lcardup_row_start', defvalue='2')
    lcardup_col_ticket = ConfigString(name='lcardup_col_ticket', defvalue='D')
    lcardup_col_resol = ConfigString(name='lcardup_col_resol', defvalue='A')
    lcardup_col_comment = ConfigString(name='lcardup_col_comment', defvalue='B')
    lcardup_status_open = ConfigBool(name='lcupstatusopen', defvalue=True)
    lcardup_status_closed = ConfigBool(name='lcupstatusclosed', defvalue=False)
    lcardup_status_investigation = ConfigBool(name='lcupstatusinvestigation', defvalue=False)
    lcardup_status_fixed = ConfigBool(name='lcupstatusfixed', defvalue=False)
    lcardup_status_rejected = ConfigBool(name='lcupstatusrejected', defvalue=False)
    lcardup_status_new = ConfigBool(name='lcupstatusnew', defvalue=False)

    def __init__(self):
        self.serverpasswordshow = False

        self.uploadsheets = list()
        self.uploadsheet = None

        self.catalogs = list()
        self.downloadcatalog = None

        self.lifecardcatalog = None
        self.lifecardtestplan = None


    @PubSend('logappend')
    def LogAppend(self, msg):
        return msg

    @PubSend('init')
    def init(self):
        return None

    @PubSend('uploadsheets')
    def GetUploadSheets(self):
        self.uploadsheets[:] = []
        wkbook = openpyxl.load_workbook(self.uploadexcel, use_iterators=True)
        self.uploadsheets = wkbook.get_sheet_names()

    @PubSend('catalogs')
    def GetUploadCatalogs(self):
        self.catalogs[:] = []
        rpc = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)
        self.catalogs = rpc.listSubCatalogsExt('')

    @PubSend('catalogdeleted')
    def DeleteCatalog(self, catalog):
        rpc = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)
        self.LogAppend('Deleting Catalog %s' % catalog[0:3])
        rpc.deleteTestCatalog(catalog[0])
        self.LogAppend('Deleted Catalog %s' % catalog[0:3])
        return catalog

    @PubSend('catalogdeletedall')
    def DeleteCatalogAll(self):
        rpc = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)
        self.LogAppend('Deleting All Catalogs')
        for catalog in self.catalogs:
            self.LogAppend('Deleting Catalog %s' % catalog[0:3])
            rpc.deleteTestCatalog(catalog[0])
            self.LogAppend('Deleted Catalog %s' % catalog[0:3])
        self.LogAppend('Deleted All Catalogs')

    #@PubSend('uploadingexcel')
    def UploadExcel(self):
        th = threading.Thread(target=self.UploadingExcel)
        th.start()

    @PubSend('uploadedexcel')
    def UploadingExcel(self):
        try:
            self.LogAppend('Using Excel file %s' % self.uploadexcel)
            # Open Workbook and sheet
            self.LogAppend('Opening excel file %s' % self.uploadexcel)
            workbook = xlrd.open_workbook(self.uploadexcel)
            chosensheet = self.uploadsheet
            self.LogAppend('Choosing sheet %s' % chosensheet)
            wksheet = workbook.sheet_by_name(chosensheet)

            self.LogAppend('Parsing TestCases Fields and categories, to create subcatalogs')
            headerfields = dict()
            catcols = 0
            tcdescription = 'Test Case Fields'
            row = 0
            while True:
                try:
                    cellvalue = wksheet.cell_value(row, catcols)
                except IndexError, e:
                    if catcols >= wksheet.ncols:
                        # expected limit reached
                        break
                    raise e
                except Exception, e:
                    raise e

                if not cellvalue:
                    break
                headerfields[catcols] = cellvalue
                tcdescription += ':%s' % cellvalue
                catcols += 1

            self.LogAppend('Excel header has %d columns' % catcols)
            self.LogAppend('Header names %s' % str(headerfields).strip('[]'))

            rpcinterface = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)
            if not self.uploadusecatalog:
                self.LogAppend('Creating Catalog %s' % self.uploadcatalogname)
                rootid = rpcinterface.createTestCatalog('', self.uploadcatalogname, tcdescription)
                self.LogAppend('Created Catalog with id %s' % rootid)
            else:
                self.LogAppend('Appending to Catalog %s: id %s' % (self.uploadusecatalogname, self.uploadusecatalogid))
                # we have set the variable uploadusecatalogid
                rootid = self._model.uploadusecatalogid

            # Parse the Category and subcategory columns
            categories = dict()
            while True:
                colidx = itertools.count()
                row += 1
                try:
                    maincat = wksheet.cell_value(row, colidx.next())
                except IndexError, e:
                    if row >= wksheet.nrows:
                        maincat = None # expected limit reached
                    else:
                        raise e # unexpected ... let's see it
                except Exception, e:
                    raise e # unexpected ... let's see it

                if not maincat: # if row >= wksheet.nrows
                    self.LogAppend('End of TestCases')
                    break # end of testcases

                # Get the main category (create if needed as child of root)
                if maincat not in categories:
                    self.LogAppend('Creating SubCatalog %s' % maincat)
                    catid = rpcinterface.createTestCatalog(rootid, maincat, '')

                    categories[maincat] = catid
                    self.LogAppend('Created SubCatalog %s with id %s' % (maincat, catid))
                else:
                    self.LogAppend('Next Testcase - Category %s' % maincat)
                    catid = categories[maincat]

                # Get the sub category (create if needed as child of main category)
                subcat = wksheet.cell_value(row, colidx.next())
                if subcat:
                    subcatkey = '%s::%s' % (maincat, subcat)
                    if subcatkey not in categories:
                        self.LogAppend('Creating SubSubCatalog %s' % subcat)
                        catid = rpcinterface.createTestCatalog(catid, subcat, '')

                        categories[subcatkey] = catid
                        self.LogAppend('Created SubSubCatalog %s with id %s' % (subcat, catid))
                    else:
                        self.LogAppend('Next Testcase - Category %s' % subcatkey)
                        catid = categories[subcatkey]

                # Get the title
                tctitle = wksheet.cell_value(row, colidx.next())

                # Get Description and complete
                tcdesc = wksheet.cell_value(row, colidx.next())
                for col in xrange(colidx.next(), catcols):
                    tcdesc += '\n\n'
                    tcdesc += '== %s ==\n' % headerfields[col]
                    cellvalue = wksheet.cell_value(row, col)
                    celltype = wksheet.cell_type(row, col)
                    if celltype != xlrd.XL_CELL_TEXT:
                        cellvalue = str(cellvalue)
                    tcdesc += cellvalue
                    tcdesc += '\n\n'

                # Testcase is complete and we have a cat id
                self.LogAppend('Creating testcase %d with title %s' % (row, tctitle))
                tcid = rpcinterface.createTestCase(catid, tctitle, tcdesc)

                self.LogAppend('%d TestCase create (id %s/catalog id %s)' % (row, tcid, catid))

        except Exception, e:
            self.LogAppend('Error during excel upload: %s' % str(e))

        self.GetUploadCatalogs()


    #@PubSend('downloadingexcel')
    def DownloadExcel(self):
        th = threading.Thread(target=self.DownloadingExcel)
        th.start()

    @PubSend('downloadedexcel')
    def DownloadingExcel(self):
        try:
            self.LogAppend('Downloading Catalog to Excel')
            self.LogAppend('Preparing root catalog')
            rootcat = self.downloadcatalog
            rootid = rootcat[0]

            # Prepare for Excel
            sheetname = rootcat[2]
            if not sheetname:
                self.LogAppend('Catalog has no name, using "Test Catalog" for the Excel Tab')
                sheetname = 'Test Catalog'
            else:
                self.LogAppend('Excel Tab: %s' % sheetname)

            # Open Excel
            # wkbook = openpyxl.reader.excel.load_workbook(self.downloadexcel)
            self.LogAppend('Opening a workbook in memory')
            wkbook = openpyxl.Workbook()
            self.LogAppend('Creating the tab')
            # wksheet = wkbook.create_sheet(title=sheetname)
            wksheet = wkbook.get_active_sheet()
            wksheet.title = sheetname

            # Connect to Server
            self.LogAppend('Creating an RPC interface to the server')
            rpcinterface = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)

            # FIXME: I am getting the full catalog from the interface
            if False:
                rootcat = rpcinterface.getTestCatalog(rootid)
                rootcat.insert(0, rootid) # use to create a full hierarcha of catalogs
            self.LogAppend('Preparing root catalog and headers')
            catalogs = [rootcat]
            catdesc = rootcat[3]
            headers = catdesc.split(':')[1:]

            row = 0
            for col, header in enumerate(headers):
                wksheet.cell(row=row, column=col).value = header

            self.LogAppend('Start testcases export')
            tt2category = dict()
            # Go for it
            while catalogs:
                nextcatalogs = list()
                self.LogAppend('Iterating over %d catalogs' % len(catalogs))
                for catalog in catalogs:
                    self.LogAppend('Exploring catalog %s' % str(catalog))
                    tt2category[catalog[0]] = catalog[2] # map id to category (catalog name)

                    self.LogAppend('Getting subcatalogs of catalogs')
                    subcatalogs = rpcinterface.listSubCatalogs(catalog[0])
                    nextcatalogs.extend(subcatalogs)

                    self.LogAppend('Getting testcases of current catalog')
                    testcases = rpcinterface.listTestCases(catalog[0])
                    self.LogAppend('Iterating over %d testcases' % len(testcases))
                    for tc in testcases:
                        self.LogAppend('Exploring testcases %s' % str(tc[0:3]))
                        row += 1
                        colidx = itertools.count()
                        tctitle = tc[2]
                        tcdesc = tc[3]
                        # TC_TT0_TT1_TC32
                        # TC_TT0_TT1_TT2_TC33
                        tcname = tc[1]
                        txt = tcname[2:]
                        txt = txt.split('_TC')[0]
                        ttparts = txt.split('_TT')[1:]
                        category = ''
                        subcategory = ''

                        if len(ttparts) > 1:
                            category = tt2category[ttparts[1]]
                        if len(ttparts) > 2:
                            subcategory = tt2category[ttparts[2]]

                        self.LogAppend('Testcase in category/subcategory %s%s' % (category, subcategory))
                        self.LogAppend('Writing testcase to excel')

                        wksheet.cell(row=row, column=colidx.next()).value = category
                        wksheet.cell(row=row, column=colidx.next()).value = subcategory
                        wksheet.cell(row=row, column=colidx.next()).value = tctitle

                        self.LogAppend('Parsing testcase description')
                        tcdescparts = self.ParseDesc(tcdesc, headers)
                        self.LogAppend('Parsing testcase description got %d parts' % len(tcdescparts))
                        for desckey, descfield in tcdescparts.iteritems():
                            col = colidx.next()
                            wksheet.cell(row=row, column=col).set_value_explicit(descfield)
                            wksheet.cell(row=row, column=col).style.alignment.wrap_text = True

                self.LogAppend('End of testcases for current catalog')
                catalogs = nextcatalogs

            self.LogAppend('Saving Workbook to disk')
            wkbook.save(self.downloadexcel)
            self.LogAppend('Done Downloading to Excel')
        except Exception, e:
            self.LogAppend('Error during excel download: %s' % str(e))


    def ParseDesc(self, desc, headers, startcol=3):
        curheader = headers[startcol]
        parsed = OrderedDict([(curheader, '')])

        lines = desc.split('\n')
        for line in lines:
            if line.startswith('=='):
                # curheader = line.strip('\n= ')
                startcol += 1
                curheader = headers[startcol]
                parsed[curheader] = ''
            elif not line:
                continue
            else:
                parsed[curheader] += line + '\n'

        return parsed

    @PubSend('testplans')
    def GetTestPlans(self, catalog):
        rpc = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)
        catid = catalog[0]
        testplans = rpc.listTestPlans(catid)
        return testplans

    #@PubSend('lifecarddownloadingexcel')
    def LifeCardDownloadExcel(self):
        th = threading.Thread(target=self.LifeCardDownloadingExcel)
        th.start()

    @PubSend('lifecarddownloadedexcel')
    def LifeCardDownloadingExcel(self):
        '''
        How it works:
         -- Get a list of Testcases already in plan
         -- Get a list of the generic Testcases, since some fields will be needed
         -- For this second list make "title" indexed dictionary to correlate with
            testcases in plan
         -- Get the Catalog tree and generate a dictionariy of id -> (Category, SubCategory) labels
         -- Generate a list of lists (range or row, cols data in Excel Terms)
         -- Each row has the the 11 fields
         -- Open Excel, the file and get the sheet
         -- Calculate offsets and write the data to the range
         -- Save, Close and Quit Excel
        '''
        wkbook = None
        tc_row_offset = 2
        tc_col_offset = 1
        ticket_row_offset = 2
        ticket_col_offset = 4
        try:
            self.LogAppend('Creating an RPC interface to the server')
            rpc = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)

            # Get Test Cases In Plan - and Order Them
            self.LogAppend('Compiling list of TestCases in Plan from server')
            tcips = rpc.listTestCasesExt(self.lifecardcatalog[0], self.lifecardtestplan[0], True)
            self.LogAppend('Got %d testcases, processing them' % len(tcips))
            tcips = map(TestCaseInPlan, tcips) # build a list of TestCaseInPlan
            tcips.sort(key=attrgetter('title')) # sort the list in place using title (our test_id)

            # Get Test Cases (for additional needed fields)
            self.LogAppend('Compiling list of TestCases from server')
            tcases = rpc.listTestCasesExt(self.lifecardcatalog[0], '', True) # Get from server
            it_tcases = map(TestCase, tcases) # Create the objects
            tcases = dict(map(lambda x: (x.title, x), it_tcases)) # make a title indexed
            tcasesb = dict(map(lambda x: (x.page_name, x), it_tcases)) # make a page_name indexed dict

            # Get the Test Catalogs and build a dictionary of ids / we only have two levels
            self.LogAppend('Start processing catalogs')
            testcats = dict()
            for catalog in map(TestCatalog, rpc.listSubCatalogsExt(self.lifecardcatalog[0])):
                testcats[catalog.id] = (catalog.title, '')
                for subcat in map(TestCatalog, rpc.listSubCatalogsExt(catalog.id)):
                    testcats[subcat.id] = (catalog.title, subcat.title)
            
            self.LogAppend('Creating Excel Range Data')
            lrange = list()
            # Create a the range input for Excel (list of lists)
            for tcip in tcips:
                ltcase = list()
                tcase = tcases[tcip.title]

                ltcase.extend(testcats[tcase.get_catalog_id()]) # category - subcategory
                ltcase.extend([tcase.title, tcase.cpeid, tcase.headline])
                ltcase.extend([tcip.get_status(), tcip.comment, tcip.tracefile, tcip.testednetwork,
                               tcip.author, tcip.timestamp.split('T')[0]])
                lrange.append(ltcase)

            self.LogAppend('Compiling ticket list')
            since = datetime.datetime(year=2014, month=1, day=1)
            ticket_ids = rpc.getRecentChanges(since)
            multicall = rpc.multicall()

            self.LogAppend('Getting tickets')
            for ticket_id in ticket_ids:
                multicall.ticket.get(ticket_id)
            tickets = map(Ticket, multicall())
            tickets.sort(key=attrgetter('id'))

            self.LogAppend('Getting tickets changelog')
            multicall = rpc.multicall()
            for ticket_id in ticket_ids:
                multicall.ticket.changeLog(ticket_id)
            # create the ticketchangelog objects from call result
            tlogs = multicall()
            tlogs = dict(zip(ticket_ids, tlogs))
            for ticket_id, clogs in tlogs.iteritems():
                tloglist = list()
                for clog in clogs:
                    tloglist.append(TicketChangeLog(clog))
                tlogs[ticket_id] = tloglist

            tstatusdown = list()
            tstatusdown.append('new') if self.lcdownnew else None
            tstatusdown.append('open') if self.lcdownopen else None
            tstatusdown.append('closed') if self.lcdownclosed else None
            tstatusdown.append('fixed') if self.lcdownfixed else None
            tstatusdown.append('rejected') if self.lcdownreject else None
            tstatusdown.append('investigation') if self.lcdowninvest else None
            self.LogAppend('Will download following ticket types: %s' % str(tstatusdown))

            self.LogAppend('Generating Excel Data Range for tickets')
            ltrange = list()
            if self.lcdowncopywithvendorcomments:
                ltrange2 = list()
            for ticket in tickets:
                if ticket.status not in tstatusdown:
                    self.LogAppend('Skipping ticket %d with status %s' % (ticket.id, ticket.status))
                    continue
                lticket = list()
                lticket.extend([ticket.id, ticket.version, ticket.reporter])
                # lticket.append('%s\n%s' % (ticket.summary, ticket.description))
                lticket.append('%s\n---------------------\n%s' % (ticket.summary, ticket.description))
                if ticket.testcaseid in tcasesb:
                    lticket.append(tcasesb[ticket.testcaseid].title)
                else:
                    lticket.append('Exploratory Testing')
                lticket.append(ticket.status.capitalize())
                lticket.append(ticket.priority)
                # lticket.append('') # skip 'Test Comment Column'
                lticket.append(ticket.created)

                tlog = tlogs[ticket.id]
                owncomment = ''
                vencomment = ''
                for comment in filter(lambda x: x.name == 'comment', tlog):
                    # format timestamp author and comment
                    if not comment.new:
                        continue
                    if comment.author == self.lifecardauthorup:
                        vencomment += '[%s]\n' % (comment.tstamp.strftime('%Y-%m-%dT%H:%M'),)
                        vencomment += comment.new
                        vencomment += '\n'
                    else:
                        owncomment += '[%s] %s\n' % (comment.tstamp.strftime('%Y-%m-%dT%H:%M'), comment.author)
                        owncomment += comment.new
                        owncomment += '\n'

                lticket.append(owncomment)
                # lticket.extend(['', '', ]) # skip 2 columnts
                if not self.lcdownfiltervendorcomments:
                    lticket.append(vencomment)
                else:
                    lticket.append('')

                resoltxt = ''
                for resolution in filter(lambda x: x.name == 'resolution', tlog):
                    if not resolution.new:
                        continue
                    resoltxt += '[%s]\n' % (resolution.tstamp.strftime('%Y-%m-%dT%H:%M'),)
                    resoltxt += resolution.new
                    resoltxt += '\n'

                lticket.append(resoltxt)

                ltrange.append(lticket)
                if self.lcdowncopywithvendorcomments:
                    lticket2 = copy.deepcopy(lticket)
                    lticket2[-2] = vencomment
                    ltrange2.append(lticket2)
            # raise Exception('Ticket List Compiled')

            if not lrange and not ltrange:
                self.LogAppend('Nothing to download to excel')
            else:

                if self.lcdownmakecopy:
                    curdate = datetime.date.today()
                    prepdate = curdate.strftime('%Y-%m-%d')
                    appcw = int(curdate.strftime('%W'))
                    curdate.replace(month=1, day=1)
                    if curdate.weekday != 0:
                        # Jan 1st is not monday, so all days until
                        # first monday are isoweek = 0 and we want
                        # calendar week 1
                        appcw += 1
                    appcw = 'cw%02d' % appcw
                    dirname, basename = os.path.split(self.lifecardexcel)
                    excelbase, excelext = os.path.splitext(basename)
                    basename = prepdate + '-' + excelbase + '-' + appcw + excelext
                    excelfile = os.path.join(dirname, basename)
                    self.LogAppend('Copying %s -> %s' % (self.lifecardexcel, excelfile))
                    shutil.copyfile(self.lifecardexcel, excelfile)
                else:
                    excelfile = self.lifecardexcel

                if self.lcdowncopywithvendorcomments:
                    vendirname, venbasename = os.path.split(excelfile)
                    venname, venext = os.path.splitext(venbasename)
                    venname += '-vendor'
                    venbasename = venname + venext
                    venexcelfile = os.path.join(vendirname, venbasename)
                    shutil.copyfile(excelfile, venexcelfile)

                exinstances = [(excelfile, ltrange),]
                if self.lcdowncopywithvendorcomments:
                    exinstances.append((venexcelfile, ltrange2))

                wkbooks = list()
                xls = list()

                for excel_file, ticket_range in exinstances:
                    self.LogAppend('Opening Workbook %s' % excelfile)
                    pythoncom.CoInitialize()
                    # xl = Dispatch('Excel.Application')
                    # xl = win32com.client.gencache.EnsureDispatchEx("Excel.Application")
                    xl = DispatchEx('Excel.Application') # Opens different instance
                    xls.append(xl)
                    xl.Visible = 1
                    # xl.Interactive = True if self.lcdownkeepexcelopen else False
                    xl.Interactive = True
                    wkbook = xl.Workbooks.Open(excel_file)
                    wkbooks.append(wkbook)
                    if not lrange:
                        self.LogAppend('No test cases to write down to Excel')
                    else:
                        lempty = [[''] * len(lrange[0])] * len(lrange)

                        wksheet = wkbook.Sheets('Test Cases')
                        if False and wksheet.AutoFilterMode:
                            wksheet.ShowAllData()
                        topleft = wksheet.Cells(tc_row_offset, tc_col_offset)
                        botright = wksheet.Cells(tc_row_offset + len(lrange) - 1, tc_col_offset + len(lrange[0]) - 1)
                        self.LogAppend('Writing Data to Worksheet')
                        wksheet.Range(topleft, botright).Value = lempty
                        wksheet.Range(topleft, botright).Value = lrange
                        self.LogAppend('End downloading test-case-in-plan information')

                    if not ticket_range:
                        self.LogAppend('No tickets to write down to Excel')
                    else:
                        self.LogAppend('Writing Data to Worksheet')
                        # FIXME: if there are no tickets ... I would also need to clear the list
                        # The property .UsedRange should give access to cells that have already
                        # been used (I can delete the range)
                        lempty = [['',] * len(ticket_range[0])] * len(tickets)

                        wksheet = wkbook.Sheets('Bug Tracking')
                        if False and wksheet.AutoFilterMode:
                            wksheet.ShowAllData()
                        topleft = wksheet.Cells(ticket_row_offset, ticket_col_offset)
                        botright = wksheet.Cells(ticket_row_offset + len(ticket_range) - 1,
                                                 ticket_col_offset + len(ticket_range[0]) - 1)

                        botrightempty = wksheet.Cells(ticket_row_offset + len(tickets) - 1,
                                                      ticket_col_offset + len(ticket_range[0]) - 1)

                        wksheet.Range(topleft, botrightempty).Value = lempty
                        wksheet.Range(topleft, botright).Value = ticket_range
                        self.LogAppend('End downloading tickets information to %s' % excel_file)

                    try:
                        if not self.lcdownexcelnotsave:
                            self.LogAppend('Saving workbook')
                            wkbook.Save()
                        if not self.lcdownkeepexcelopen:
                            self.LogAppend('Closing workbook and Quitting Excel')
                            wkbook.Close(False)
                            wkbooks.remove(wkbook)
                            xl.Quit()
                            xls.remove(xl)
                    except Exception, e:
                        self.LogAppend('Error saving/closing/quitting Excel file: %s' % str(e))

                self.LogAppend('End downloading to Lifecard')

        except Exception, e:
            self.LogAppend('Error during Download to LifeCard: %s' % str(e))

        if not self.lcdownkeepexcelopen:
            self.LogAppend('Trying to clean up any Excel leftover if needed')
            for wkbook in wkbooks:
                try:
                    wkbook.Close(False)
                except Exception:
                    pass

            for xl in xls:
                try:
                    xl.Quit()
                except Exception:
                    pass

        self.LogAppend('End cleaning up to Lifecard')


    #@PubSend('lifecarddownloadingattach')
    def LifeCardDownloadAttach(self):
        th = threading.Thread(target=self.LifeCardDownloadingAttach)
        th.start()

    @PubSend('lifecarddownloadedattach')
    def LifeCardDownloadingAttach(self):
        '''
        How it works:
          -- TBD
        '''
        try:
            self.LogAppend('Downloading the attachments')
            rpc = RpcInterface(self.serverurl, self.serverusername, self.serverpassword)

            self.LogAppend('Compiling ticket list')
            since = datetime.datetime(year=2014, month=1, day=1)
            ticket_ids = rpc.getRecentChanges(since)
            multicall = rpc.multicall()
            for ticket_id in ticket_ids:
                multicall.ticket.get(ticket_id)

            self.LogAppend('Getting tickets')
            tickets = map(Ticket, multicall())
            tickets.sort(key=attrgetter('id'))
            self.LogAppend('Got %d ticket' % len(tickets))

            tstatusdown = list()
            tstatusdown.append('new') if self.lcdownattachnew else None
            tstatusdown.append('open') if self.lcdownattachopen else None
            tstatusdown.append('closed') if self.lcdownattachclosed else None
            tstatusdown.append('fixed') if self.lcdownattachfixed else None
            tstatusdown.append('rejected') if self.lcdownattachrejected else None
            tstatusdown.append('investigation') if self.lcdownattachinvestigation else None
            self.LogAppend('Will download attachment for following ticket types: %s' % str(tstatusdown))

            self.LogAppend('Processing attachments')
            # Time to retrieve and save
            for ticket in tickets:
                if ticket.status not in tstatusdown:
                    self.LogAppend('Skipping ticket %d with status %s' % (ticket.id, ticket.status))
                    continue

                self.LogAppend('Getting list of attachments ticket for ticket %d' % ticket.id)
                ticket.attachments = rpc.listAttachments(ticket.id)
                self.LogAppend('Got %d attachments ticket for ticket %d' % (len(ticket.attachments), ticket.id))
                if ticket.attachments:
                    self.LogAppend('Downloading ticket %d attachments' % ticket.id)

                for attach in ticket.attachments:
                    try:
                        battach = rpc.getAttachment(ticket.id, attach[0])
                    except Exception, e:
                        self.LogAppend('Ticket %d Attachment Missing %s: %s' % (ticket.id, attach[0], str(e)))
                    dirname = self.lifecardattachdir + '\\' + str(ticket.id).zfill(3)
                    if not os.path.isdir(dirname):
                        try:
                            self.LogAppend('Creating dir %s' % dirname)
                            os.makedirs(dirname)
                        except os.error, e:
                            raise Exception('Could not create dir %s: %s' % (dirname, str(e)))

                    try:
                        ofile = open(dirname + '\\' + attach[0], "wb")
                        ofile.write(battach.data)
                        ofile.close
                    except Exception, e:
                        raise Exception('Could not write file %s in dir %s: %s' % (attach[0], dirname, str(e)))

            self.LogAppend('End downloading Attachments')

        except Exception, e:
            self.LogAppend('Error saving attachments: %s' % str(e))


    #@PubSend('lifecarduploading')
    def LifeCardUpload(self):
        th = threading.Thread(target=self.LifeCardUploading)
        th.start()

    @PubSend('lifecarduploaded')
    def LifeCardUploading(self):
        '''
        How it works:
          -- TBD
        '''

        try:
            self.LogAppend('Getting updates from %s/%s' % (self.lifecardexcelup, self.lcardup_sheetname))
            exinput = ExcelInput(self.lifecardexcelup)
            updates = exinput.get_lifecard_updates(
                self.lcardup_sheetname, self.lcardup_row_start,
                self.lcardup_col_ticket, self.lcardup_col_resol, self.lcardup_col_comment)

            if not updates:
                self.LogAppend('No updated found on Lifecard')
                return

            self.LogAppend('%d updates found' % len(updates))

            self.LogAppend('Opening interface to the server')
            rpc = RpcSupport(self.serverurl, self.serverusername, self.serverpassword)
            self.LogAppend('Compiling ticket list from server')
            tickets = rpc.get_tickets_by_id(updates.keys())


            self.LogAppend('Updating tickets')
            tstatusup = list()
            tstatusup.append('open') if self.lcardup_status_open else None
            tstatusup.append('closed') if self.lcardup_status_closed else None
            tstatusup.append('investigation') if self.lcardup_status_investigation else None
            tstatusup.append('fixed') if self.lcardup_status_fixed else None
            tstatusup.append('rejected') if self.lcardup_status_rejected else None
            tstatusup.append('new') if self.lcardup_status_new else None
            self.LogAppend('Will upload following ticket types: %s' % str(tstatusup))

            self.LogAppend('Updating tickets')
            ticket_updates = list()
            for ticket_id, value in updates.iteritems():
                resolution, comment = value
                ticket = tickets[ticket_id]
                if ticket.status not in tstatusup:
                    self.LogAppend('Skipping Ticket %d with status (%s)' % (ticket.id, ticket.status))
                    continue
                self.LogAppend('Updating Ticket %d' % ticket_id)
                attributes = dict()
                if resolution:
                    attributes['resolution'] = resolution

                # Old Update --- Check ticket.py code and look for an example of 'action'
                # ticket_id, comment, attributes, notify, author, when are the parameters
                ticket_updates.append((ticket_id, comment, attributes, False, self.lifecardauthorup, None))

            self.LogAppend('Sending ticket updates to server')
            rpc.update_tickets(ticket_updates)

            self.LogAppend('Finished updating Lifecard')

        except Exception, e:
            self.LogAppend('Error during LifeCard upload: %s' % str(e))
